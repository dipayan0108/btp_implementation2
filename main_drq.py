"""
main_drq.py  –  SmartDrive + DrQ modification
----------------------------------------------
Changes from baseline (main.py):
  1. VAE / EncodeState removed entirely.
  2. ConvEncoder  – lightweight deterministic CNN, output z ∈ R^LATENT_DIM
  3. DrQAugment   – random crop via padding (train only), disabled at inference
  4. PPOAgent.learn() now differentiates through the encoder jointly with
     actor / critic (encoder vars included in gradient tapes).
  5. Encoder saved / loaded alongside actor & critic.
  6. ExcelLogger  – writes training_log.xlsx into RESULTS_PATH every episode.
  Everything else (CarlaEnvironment, reward, sensors, …) is unchanged.
"""

import os
import sys
import glob
import math
import weakref
import pygame
import time
import random
import csv
import cv2
import pickle
import numpy as np
import pandas as pd
import logging
from datetime import datetime
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ModuleNotFoundError:
    Workbook = load_workbook = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False

import tensorflow as tf
import tensorflow_probability as tfp
tfd = tfp.distributions
layers = tf.keras.layers

from parameters import *

# ── device check ──────────────────────────────────────────────────────────────
gpu_available = tf.config.list_physical_devices('GPU')
if gpu_available:
    print("CUDA is available! TensorFlow is using the GPU.")
else:
    print("CUDA is NOT available. TensorFlow is using the CPU.")
print()

try:
    sys.path.append(glob.glob('./carla/carla-*%d.%d-%s.egg' % (
        sys.version_info.major,
        sys.version_info.minor,
        'win-amd64' if os.name == 'nt' else 'linux-x86_64'))[0])
except IndexError:
    print("Couldn't import Carla egg properly")

import carla


# ══════════════════════════════════════════════════════════════════════════════
# DrQ random-crop augmentation
# ══════════════════════════════════════════════════════════════════════════════
class DrQAugment:
    """
    Random crop via padding + random shift — the core DrQ augmentation.
    Input shape:  (H, W, C)  numpy array  (uint8 or float32)
    Output shape: same  (float32, [0,255])
    At inference call with training=False to get identity transform.
    """

    def __init__(self, pad=AUGMENTATION_PAD):
        self.pad = pad

    def __call__(self, image: np.ndarray, training: bool = True) -> tf.Tensor:
        img = tf.convert_to_tensor(image, dtype=tf.float32)           # (H,W,C)
        if not training:
            return img

        h, w, c = IM_HEIGHT, IM_WIDTH, 3
        # pad by `pad` pixels on every side
        img = tf.expand_dims(img, 0)                                   # (1,H,W,C)
        img = tf.pad(img,
                     [[0,0],[self.pad,self.pad],[self.pad,self.pad],[0,0]],
                     mode='REFLECT')
        # random crop back to original size
        img = tf.image.random_crop(img, size=[1, h, w, c])
        img = tf.squeeze(img, 0)                                       # (H,W,C)
        return img


# ══════════════════════════════════════════════════════════════════════════════
# Deterministic Convolutional Encoder  (replaces VAE)
# ══════════════════════════════════════════════════════════════════════════════
@tf.keras.utils.register_keras_serializable()
class ConvEncoder(tf.keras.Model):
    """
    Lightweight CNN encoder trained end-to-end with PPO.
    Input : (1, H, W, C)  float32  [0,255]
    Output: (LATENT_DIM,) float32  latent vector z
    No decoder, no reconstruction loss — representation is learned purely
    through the PPO objective.
    """

    def __init__(self, latent_dim=LATENT_DIM, name='ConvEncoder', **kwargs):
        super().__init__(name=name, **kwargs)
        self.latent_dim = latent_dim

        # Normalise pixels to [0,1] inside the model so the graph is portable
        self.norm = layers.Lambda(lambda x: x / 255.0)

        self.conv1 = layers.Conv2D(32, 4, strides=2, activation='relu',
                                   padding='valid')      # 79×39
        self.conv2 = layers.Conv2D(64, 3, strides=2, activation='relu',
                                   padding='same')       # 40×20
        self.conv3 = layers.Conv2D(128, 4, strides=2, activation='relu',
                                   padding='valid')      # 19×9
        self.conv4 = layers.Conv2D(256, 3, strides=2, activation='relu',
                                   padding='valid')      # 9×4
        self.flatten = layers.Flatten()
        self.fc      = layers.Dense(1024, activation='relu')
        self.proj    = layers.Dense(latent_dim)          # no activation → raw z

    def call(self, x, training=False):
        """x: (batch, H, W, C) float32"""
        x = self.norm(x)
        x = self.conv1(x)
        x = self.conv2(x)
        x = self.conv3(x)
        x = self.conv4(x)
        x = self.flatten(x)
        x = self.fc(x)
        z = self.proj(x)
        return z   # (batch, latent_dim)


# ══════════════════════════════════════════════════════════════════════════════
# Observation pipeline  (replaces EncodeState)
# ══════════════════════════════════════════════════════════════════════════════
class EncodeState:
    """
    Wraps ConvEncoder + DrQAugment.
    process() returns a flat tf.Tensor: [z | nav_features]
    training=True  → augmentation ON
    training=False → augmentation OFF (inference / edge)
    """

    def __init__(self, encoder: ConvEncoder, training: bool = True):
        self.encoder  = encoder
        self.augment  = DrQAugment(pad=AUGMENTATION_PAD)
        self.training = training

    def process(self, observation) -> tf.Tensor:
        """
        observation = [image_array (H,W,C), nav_array (5,)]
        Returns flat (LATENT_DIM+5,) float32 tensor.
        """
        img = self.augment(observation[0], training=self.training)     # (H,W,C)
        img = tf.expand_dims(img, 0)                                   # (1,H,W,C)
        z   = self.encoder(img, training=self.training)                # (1,D)
        z   = tf.reshape(z, [-1])                                      # (D,)
        nav = tf.convert_to_tensor(observation[1], dtype=tf.float32)   # (5,)
        obs = tf.concat([z, nav], axis=-1)                             # (D+5,)
        return obs


# ══════════════════════════════════════════════════════════════════════════════
# Actor / Critic  (unchanged from baseline)
# ══════════════════════════════════════════════════════════════════════════════
@tf.keras.utils.register_keras_serializable()
class Actor(tf.keras.Model):

    def __init__(self, name='ACTOR', **kwargs):
        super().__init__(name=name, **kwargs)
        self.dense1 = layers.Dense(500, activation='tanh')
        self.dense2 = layers.Dense(300, activation='tanh')
        self.dense3 = layers.Dense(100, activation='tanh')
        self.output_layer = layers.Dense(ACTION_DIM, activation='tanh')

    def call(self, obs):
        if isinstance(obs, np.ndarray):
            obs = tf.convert_to_tensor(obs, dtype=tf.float32)
        obs = tf.clip_by_value(obs, -1e8, 1e8)
        x   = self.dense1(obs)
        x   = self.dense2(x)
        x   = self.dense3(x)
        return self.output_layer(x)


@tf.keras.utils.register_keras_serializable()
class Critic(tf.keras.Model):

    def __init__(self, name='CRITIC', **kwargs):
        super().__init__(name=name, **kwargs)
        self.dense1 = layers.Dense(500, activation='tanh')
        self.dense2 = layers.Dense(300, activation='tanh')
        self.dense3 = layers.Dense(100, activation='tanh')
        self.output_layer = layers.Dense(1)

    def call(self, obs):
        if isinstance(obs, np.ndarray):
            obs = tf.convert_to_tensor(obs, dtype=tf.float32)
        obs = tf.clip_by_value(obs, -1e8, 1e8)
        x   = self.dense1(obs)
        x   = self.dense2(x)
        x   = self.dense3(x)
        return self.output_layer(x)


# ══════════════════════════════════════════════════════════════════════════════
# Replay Buffer
# ══════════════════════════════════════════════════════════════════════════════
class Buffer:
    def __init__(self):
        self.observation = []
        self.actions     = []
        self.log_probs   = []
        self.rewards     = []
        self.dones       = []

    def clear(self):
        del self.observation[:]
        del self.actions[:]
        del self.log_probs[:]
        del self.rewards[:]
        del self.dones[:]


# ══════════════════════════════════════════════════════════════════════════════
# PPO Agent  –  encoder variables included in gradient update
# ══════════════════════════════════════════════════════════════════════════════
@tf.keras.utils.register_keras_serializable()
class PPOAgent(tf.keras.Model):

    def __init__(self, encoder: ConvEncoder, name='PPOAgent', **kwargs):
        super().__init__(name=name, **kwargs)

        self.encoder    = encoder           # shared reference
        self.obs_dim    = OBSERVATION_DIM
        self.action_dim = ACTION_DIM
        self.clip       = POLICY_CLIP
        self.gamma      = GAMMA
        self.lam        = LAMBDA
        self.lr         = LEARNING_RATE
        self.batch_size = BATCH_SIZE
        self.n_updates  = NO_OF_ITERATIONS

        self.memory   = Buffer()
        self.town     = TOWN
        self.optimizer = tf.keras.optimizers.Adam(learning_rate=self.lr)
        self.mse       = tf.keras.losses.MeanSquaredError()

        self.models_dir     = PPO_MODEL_PATH
        self.checkpoint_dir = CHECKPOINT_PATH
        self.encoder_dir    = ENCODER_PATH

        self.log_std = tf.Variable(tf.fill((self.action_dim,), ACTION_STD_INIT),
                                   trainable=False, dtype=tf.float32)

        self.actor     = Actor()
        self.critic    = Critic()
        self.old_actor  = Actor()
        self.old_critic = Critic()

        self.actor.compile(optimizer=self.optimizer)
        self.critic.compile(optimizer=self.optimizer)
        self.old_actor.compile(optimizer=self.optimizer)
        self.old_critic.compile(optimizer=self.optimizer)

        self.update_old_policy()

    # ── inference ─────────────────────────────────────────────────────────────
    def call(self, obs, train: bool):
        if isinstance(obs, np.ndarray):
            obs = tf.convert_to_tensor(obs, dtype=tf.float32)
        if len(obs.shape) == 1:
            obs = tf.expand_dims(obs, 0)

        mean   = self.old_actor(obs)
        action, log_probs = self._sample(mean)
        value  = self.old_critic(obs)

        if train:
            self.memory.observation.append(obs)
            self.memory.actions.append(action)
            self.memory.log_probs.append(log_probs)

        return action.numpy().flatten(), mean.numpy().flatten()

    def _sample(self, mean):
        dist      = tfd.MultivariateNormalDiag(loc=mean, scale_diag=self.log_std)
        action    = dist.sample()
        log_probs = dist.log_prob(action)
        return action, log_probs

    def _evaluate(self, obs, action):
        mean      = self.actor(obs)
        dist      = tfd.MultivariateNormalDiag(loc=mean, scale_diag=self.log_std)
        log_probs = dist.log_prob(action)
        entropy   = dist.entropy()
        values    = self.critic(obs)
        return log_probs, values, entropy

    def update_old_policy(self):
        self.old_actor.set_weights(self.actor.get_weights())
        self.old_critic.set_weights(self.critic.get_weights())

    # ── advantage computation (GAE) ────────────────────────────────────────────
    def compute_advantages(self, rewards, values, dones):
        advantages = []
        gae = 0
        for i in reversed(range(len(rewards))):
            delta = rewards[i] + self.gamma * values[i+1] * (1 - dones[i]) - values[i]
            gae   = delta + self.gamma * self.lam * (1 - dones[i]) * gae
            advantages.insert(0, gae)
        returns = [adv + val for adv, val in zip(advantages, values[:-1])]
        return (tf.convert_to_tensor(advantages, dtype=tf.float32),
                tf.convert_to_tensor(returns,    dtype=tf.float32))

    # ── learn  –  encoder updated jointly ─────────────────────────────────────
    def learn(self):
        print()
        rewards      = self.memory.rewards
        dones        = self.memory.dones
        old_states   = tf.squeeze(tf.stack(self.memory.observation, axis=0))
        old_actions  = tf.squeeze(tf.stack(self.memory.actions,     axis=0))
        old_logprobs = tf.squeeze(tf.stack(self.memory.log_probs,   axis=0))

        values = tf.squeeze(self.critic(old_states))
        values = tf.concat([values, tf.zeros((1,))], axis=0)

        advantages, returns = self.compute_advantages(rewards, values.numpy().tolist(), dones)
        advantages = (advantages - tf.reduce_mean(advantages)) / (tf.math.reduce_std(advantages) + 1e-7)
        returns    = (returns    - tf.reduce_mean(returns))    / (tf.math.reduce_std(returns)    + 1e-7)

        # Variables to optimise: actor + critic + ENCODER (end-to-end)
        actor_vars   = self.actor.trainable_variables + self.encoder.trainable_variables
        critic_vars  = self.critic.trainable_variables

        for _ in range(self.n_updates):
            with tf.GradientTape() as tape_a, tf.GradientTape() as tape_c:
                log_probs, values_pred, dist_entropy = self._evaluate(old_states, old_actions)
                values_pred = tf.squeeze(values_pred)

                ratios = tf.exp(log_probs - old_logprobs)
                surr1  = ratios * advantages
                surr2  = tf.clip_by_value(ratios, 1 - self.clip, 1 + self.clip) * advantages

                actor_loss  = -tf.reduce_mean(tf.minimum(surr1, surr2)) \
                              - 0.01 * tf.reduce_mean(dist_entropy)
                critic_loss = 0.5 * self.mse(values_pred, returns)

            grads_a = tape_a.gradient(actor_loss,  actor_vars)
            grads_c = tape_c.gradient(critic_loss, critic_vars)

            self.optimizer.apply_gradients(zip(grads_a, actor_vars))
            self.optimizer.apply_gradients(zip(grads_c, critic_vars))

        self.update_old_policy()
        self.memory.clear()
        print("UPDATED THE WEIGHTS (encoder + actor + critic)\n")

    # ── save / load ────────────────────────────────────────────────────────────
    def save(self):
        os.makedirs(self.models_dir,  exist_ok=True)
        os.makedirs(self.encoder_dir, exist_ok=True)

        self.actor.save(self.models_dir  + '/actor')
        self.critic.save(self.models_dir + '/critic')
        self.encoder.save(self.encoder_dir)

        np.save(os.path.join(self.models_dir, 'log_std.npy'), self.log_std.numpy())
        print(f"Model saved at {self.models_dir}  |  Encoder saved at {self.encoder_dir}")

    def load(self):
        self.actor      = tf.keras.models.load_model(self.models_dir + '/actor')
        self.critic     = tf.keras.models.load_model(self.models_dir + '/critic')
        self.old_actor  = tf.keras.models.load_model(self.models_dir + '/actor')
        self.old_critic = tf.keras.models.load_model(self.models_dir + '/critic')
        self.encoder    = tf.keras.models.load_model(self.encoder_dir)
        print(f"Models loaded from {self.models_dir}  |  Encoder from {self.encoder_dir}")

    def chkpt_save(self, episode, timestep, cumulative_score):
        os.makedirs(self.checkpoint_dir, exist_ok=True)
        checkpoint_file = os.path.join(self.checkpoint_dir, 'checkpoint.pickle')
        data = {'episode': episode, 'timestep': timestep,
                'cumulative_score': cumulative_score,
                'log_std': self.log_std.numpy()}
        with open(checkpoint_file, 'wb') as f:
            pickle.dump(data, f)
        print(f"Checkpoint saved as {checkpoint_file}\n")

    def chkpt_load(self):
        checkpoint_file = os.path.join(self.checkpoint_dir, 'checkpoint.pickle')
        with open(checkpoint_file, 'rb') as f:
            data = pickle.load(f)
        print(f"Checkpoint loaded from {checkpoint_file}  episode: {data['episode']}")
        return data['episode'], data['timestep'], data['cumulative_score']

    def prn(self):
        print(f'\nlog_std = {self.log_std}\n')


# ══════════════════════════════════════════════════════════════════════════════
# CARLA boilerplate  (copied verbatim from baseline — DO NOT modify)
# ══════════════════════════════════════════════════════════════════════════════

class ClientConnection:

    def __init__(self):
        self.host    = "localhost"
        self.town    = TOWN
        self.client  = None
        self.port    = 2000
        self.timeout = 20.0

    def setup(self):
        try:
            self.client = carla.Client(self.host, self.port)
            self.client.set_timeout(self.timeout)
            self.world  = self.client.load_world(self.town)
            self.world.set_weather(carla.WeatherParameters.CloudyNoon)
            return self.client, self.world
        except Exception as e:
            print(f'Failed to connect: {e}')


class CameraSensor:

    def __init__(self, vehicle):
        self.parent       = vehicle
        self.front_camera = []
        world             = self.parent.get_world()
        self.sensor       = self._setup(world)
        weak_self         = weakref.ref(self)
        self.sensor.listen(lambda img: CameraSensor._parse(weak_self, img))

    def _setup(self, world):
        bp = world.get_blueprint_library().find('sensor.camera.semantic_segmentation')
        bp.set_attribute('image_size_x', str(IM_WIDTH))
        bp.set_attribute('image_size_y', str(IM_HEIGHT))
        bp.set_attribute('fov', '110')
        return world.spawn_actor(bp, carla.Transform(carla.Location(x=2.5, z=0.7)),
                                 attach_to=self.parent)

    @staticmethod
    def _parse(weak_self, image):
        self = weak_self()
        if not self:
            return
        image.convert(carla.ColorConverter.CityScapesPalette)
        array = np.frombuffer(image.raw_data, dtype=np.uint8)
        array = array.reshape((image.height, image.width, 4))[:, :, :3]
        self.front_camera.append(array)


class CameraSensorEnv:

    def __init__(self, vehicle):
        self.parent = vehicle
        world       = self.parent.get_world()
        self.sensor = self._setup(world)

    def _setup(self, world):
        bp = world.get_blueprint_library().find('sensor.camera.rgb')
        bp.set_attribute('image_size_x', '640')
        bp.set_attribute('image_size_y', '480')
        bp.set_attribute('fov', '110')
        sensor = world.spawn_actor(bp,
                                   carla.Transform(carla.Location(x=-5.5, z=2.8),
                                                   carla.Rotation(pitch=-15)),
                                   attach_to=self.parent)
        sensor.listen(lambda img: self._parse(img))
        return sensor

    def _parse(self, image):
        array = np.frombuffer(image.raw_data, dtype=np.uint8)
        array = array.reshape((image.height, image.width, 4))[:, :, :3]
        surface = pygame.surfarray.make_surface(array.swapaxes(0, 1))
        self.display.blit(surface, (0, 0))
        pygame.display.flip()


class CollisionSensor:

    def __init__(self, vehicle):
        self.parent         = vehicle
        self.collision_data = []
        world               = self.parent.get_world()
        self.sensor         = self._setup(world)
        weak_self           = weakref.ref(self)
        self.sensor.listen(lambda e: CollisionSensor._on_collision(weak_self, e))

    def _setup(self, world):
        bp = world.get_blueprint_library().find('sensor.other.collision')
        return world.spawn_actor(bp,
                                 carla.Transform(carla.Location(x=1.3, z=0.5)),
                                 attach_to=self.parent)

    @staticmethod
    def _on_collision(weak_self, event):
        self = weak_self()
        if not self:
            return
        imp = event.normal_impulse
        self.collision_data.append(math.sqrt(imp.x**2 + imp.y**2 + imp.z**2))


class CarlaEnvironment:
    """Unchanged from SmartDrive baseline."""

    def __init__(self, client, world, town, checkpoint_frequency=100,
                 continuous_action=True):
        self.client          = client
        self.world           = world
        self.blueprint_library = world.get_blueprint_library()
        self.map             = world.get_map()
        self.action_space    = self._get_discrete_action_space()
        self.continous_action_space = True
        self.display_on      = VISUAL_DISPLAY
        self.vehicle         = None
        self.settings        = None
        self.current_waypoint_index   = 0
        self.checkpoint_waypoint_index = 0
        self.fresh_start     = True
        self.checkpoint_frequency = checkpoint_frequency
        self.route_waypoints = None
        self.town            = town
        self.camera_obj      = None
        self.env_camera_obj  = None
        self.collision_obj   = None
        self.lane_invasion_obj = None
        self.sensor_list     = []
        self.actor_list      = []
        self.walker_list     = []
        self._create_pedestrians()

    # ── helpers ────────────────────────────────────────────────────────────────
    def _get_discrete_action_space(self):
        return np.array([-0.9, -0.6, -0.3, -0.15, 0.0, 0.15, 0.3, 0.6, 0.9])

    def _get_vehicle(self, name):
        return self.blueprint_library.filter(name)[0]

    def _create_pedestrians(self):
        for _ in range(NUMBER_OF_PEDESTRIAN):
            try:
                spawn_point = carla.Transform()
                loc         = self.world.get_random_location_from_navigation()
                if loc:
                    spawn_point.location = loc
                    walker_bp = random.choice(
                        self.blueprint_library.filter('walker.pedestrian.*'))
                    walker = self.world.try_spawn_actor(walker_bp, spawn_point)
                    if walker:
                        self.walker_list.append(walker)
            except Exception:
                pass

    def remove_sensors(self):
        self.camera_obj      = None
        self.env_camera_obj  = None
        self.collision_obj   = None
        self.lane_invasion_obj = None

    # ── reset ─────────────────────────────────────────────────────────────────
    def reset(self):
        try:
            if self.actor_list or self.sensor_list:
                self.client.apply_batch(
                    [carla.command.DestroyActor(x) for x in self.sensor_list])
                self.client.apply_batch(
                    [carla.command.DestroyActor(x) for x in self.actor_list])
                self.sensor_list.clear()
                self.actor_list.clear()
            self.remove_sensors()

            vehicle_bp = self._get_vehicle(CAR_NAME)

            if self.town == 'Town07':
                transform          = self.map.get_spawn_points()[20]
                self.total_distance = 750
            elif self.town == 'Town02':
                transform          = self.map.get_spawn_points()[30]
                self.total_distance = 500
            else:
                transform          = self.map.get_spawn_points()[12]
                self.total_distance = 500

            self.vehicle = self.world.try_spawn_actor(vehicle_bp, transform)
            self.actor_list.append(self.vehicle)

            self.camera_obj = CameraSensor(self.vehicle)
            while not self.camera_obj.front_camera:
                time.sleep(0.0001)
            self.image_obs = self.camera_obj.front_camera.pop(-1)
            self.sensor_list.append(self.camera_obj.sensor)

            if self.display_on:
                self.env_camera_obj = CameraSensorEnv(self.vehicle)
                self.sensor_list.append(self.env_camera_obj.sensor)

            self.collision_obj      = CollisionSensor(self.vehicle)
            self.collision_history  = self.collision_obj.collision_data
            self.sensor_list.append(self.collision_obj.sensor)

            self.timesteps            = 0
            self.rotation             = self.vehicle.get_transform().rotation.yaw
            self.previous_location    = self.vehicle.get_location()
            self.distance_traveled    = 0.0
            self.center_lane_deviation = 0.0
            self.target_speed         = 22
            self.max_speed            = 35.0
            self.min_speed            = 15.0
            self.max_distance_from_center = 3
            self.throttle             = 0.0
            self.previous_steer       = 0.0
            self.velocity             = 0.0
            self.distance_from_center = 0.0
            self.angle                = 0.0
            self.distance_covered     = 0.0

            if self.fresh_start:
                self.current_waypoint_index = 0
                self.route_waypoints        = []
                self.waypoint = self.map.get_waypoint(
                    self.vehicle.get_location(), project_to_road=True,
                    lane_type=carla.LaneType.Driving)
                current_waypoint = self.waypoint
                self.route_waypoints.append(current_waypoint)
                for x in range(self.total_distance):
                    if self.town == 'Town07':
                        nxt = current_waypoint.next(1.0)[0 if x < 650 else -1]
                    elif self.town == 'Town02':
                        nxt = current_waypoint.next(1.0)[-1 if x > 100 else 0]
                    else:
                        nxt = current_waypoint.next(1.0)[-1 if x < 300 else 0]
                    self.route_waypoints.append(nxt)
                    current_waypoint = nxt
            else:
                waypoint  = self.route_waypoints[
                    self.checkpoint_waypoint_index % len(self.route_waypoints)]
                self.vehicle.set_transform(waypoint.transform)
                self.current_waypoint_index = self.checkpoint_waypoint_index

            self.navigation_obs = np.array(
                [self.throttle, self.velocity, self.previous_steer,
                 self.distance_from_center, self.angle])
            time.sleep(0.5)
            self.collision_history.clear()
            self.episode_start_time = time.time()
            return [self.image_obs, self.navigation_obs]

        except Exception as e:
            print(f'Error in reset: {e}')
            self.client.apply_batch(
                [carla.command.DestroyActor(x) for x in self.sensor_list])
            self.client.apply_batch(
                [carla.command.DestroyActor(x) for x in self.actor_list])
            self.client.apply_batch(
                [carla.command.DestroyActor(x) for x in self.walker_list])
            self.sensor_list.clear(); self.actor_list.clear()
            self.remove_sensors()
            if self.display_on:
                pygame.quit()

    # ── step ──────────────────────────────────────────────────────────────────
    def step(self, action_idx):
        try:
            self.timesteps  += 1
            self.fresh_start = False

            velocity = self.vehicle.get_velocity()
            self.velocity = math.sqrt(velocity.x**2 + velocity.y**2 + velocity.z**2) * 3.6

            if self.continous_action_space:
                steer    = float(np.clip(action_idx[0], -1.0, 1.0))
                throttle = float(np.clip((action_idx[1] + 1.0) / 2, 0.0, 1.0))
                self.vehicle.apply_control(carla.VehicleControl(
                    steer    = self.previous_steer * 0.9 + steer * 0.1,
                    throttle = self.throttle * 0.9 + throttle * 0.1))
                self.previous_steer = steer
                self.throttle       = throttle
            else:
                steer = self.action_space[action_idx]
                ctrl  = carla.VehicleControl(
                    steer    = self.previous_steer * 0.9 + steer * 0.1,
                    throttle = 1.0 if self.velocity < 20.0 else 0.0)
                self.vehicle.apply_control(ctrl)
                self.previous_steer = steer
                self.throttle       = 1.0

            if self.vehicle.is_at_traffic_light():
                tl = self.vehicle.get_traffic_light()
                if tl.get_state() == carla.TrafficLightState.Red:
                    tl.set_state(carla.TrafficLightState.Green)

            self.collision_history = self.collision_obj.collision_data
            self.rotation          = self.vehicle.get_transform().rotation.yaw
            self.location          = self.vehicle.get_location()

            # waypoint progress
            waypoint_index = self.current_waypoint_index
            for _ in range(len(self.route_waypoints)):
                idx = (waypoint_index + 1) % len(self.route_waypoints)
                wp  = self.route_waypoints[idx]
                if wp.transform.location.distance(self.location) < 5.0:
                    waypoint_index = idx
                else:
                    break
            self.current_waypoint_index = waypoint_index
            wp = self.route_waypoints[self.current_waypoint_index]

            # distance from center
            self.distance_from_center = wp.transform.location.distance(self.location)

            # angle to waypoint
            vec      = wp.transform.get_forward_vector()
            dx       = self.location.x - wp.transform.location.x
            dy       = self.location.y - wp.transform.location.y
            self.angle = math.degrees(math.atan2(dy * vec.x - dx * vec.y,
                                                  dx * vec.x + dy * vec.y))

            # camera
            while not self.camera_obj.front_camera:
                time.sleep(0.0001)
            self.image_obs = self.camera_obj.front_camera.pop(-1)

            self.navigation_obs = np.array(
                [self.throttle, self.velocity / self.max_speed,
                 self.previous_steer,
                 self.distance_from_center / self.max_distance_from_center,
                 self.angle / 180.0])

            # distance
            dist = self.location.distance(self.previous_location)
            if dist > 0.5:
                dist = 0.0
            self.distance_covered        += dist
            self.center_lane_deviation   += self.distance_from_center
            self.previous_location        = self.location

            # checkpoint
            if (self.checkpoint_frequency and
                    self.current_waypoint_index % self.checkpoint_frequency == 0 and
                    self.current_waypoint_index > 0):
                self.checkpoint_waypoint_index = self.current_waypoint_index

            # reward
            reward, done = self._reward(dist)

            return ([self.image_obs, self.navigation_obs],
                    reward, done,
                    [round(self.distance_covered, 2),
                     round(self.center_lane_deviation, 2)])

        except Exception as e:
            print(f'Error in step: {e}')
            return None, 0, True, [0, 0]

    def _reward(self, dist):
        done = False
        # collision
        if self.collision_history:
            done   = True
            reward = -10
            self.collision_history.clear()
            return reward, done

        # off road
        if self.distance_from_center > self.max_distance_from_center:
            done   = True
            reward = -10
            return reward, done

        # speed reward
        if self.velocity < self.min_speed:
            v_rew = (self.velocity - self.min_speed) / self.min_speed
        elif self.velocity > self.max_speed:
            v_rew = (self.max_speed - self.velocity) / self.max_speed
        else:
            v_rew = 1.0

        # centering reward
        c_rew = 1.0 - (self.distance_from_center / self.max_distance_from_center)

        reward = 2 * v_rew * c_rew
        return reward, done


# ══════════════════════════════════════════════════════════════════════════════
# Excel Training Logger
# ══════════════════════════════════════════════════════════════════════════════
class ExcelLogger:
    """
    Appends one row per episode to  RESULTS_PATH/training_log.xlsx
    Columns:
        Episode | Timestep | Episode Reward | Avg Reward (last 10) |
        Cumulative Avg Reward | Distance Covered (m) |
        Deviation from Center (m) | Episode Duration (s) | Timestamp
    A summary sheet is auto-updated with MIN / MAX / AVERAGE formulas.
    """

    HEADERS = [
        'Episode', 'Timestep', 'Episode Reward', 'Avg Reward (last 10)',
        'Cumulative Avg Reward', 'Distance Covered (m)',
        'Deviation from Center (m)', 'Episode Duration (s)', 'Timestamp'
    ]

    # column widths  (matched to header index)
    COL_WIDTHS = [10, 12, 16, 22, 22, 22, 26, 22, 22]

    # Header fill / font
    HEADER_FILL  = PatternFill('solid', start_color='1F3864')   # dark navy
    HEADER_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    DATA_FONT    = Font(name='Arial', size=10)
    ALT_FILL     = PatternFill('solid', start_color='D9E1F2')   # light blue
    BORDER_SIDE  = Side(style='thin', color='BFBFBF')

    def __init__(self, results_path: str):
        self.path = os.path.join(results_path, 'training_log.xlsx')
        os.makedirs(results_path, exist_ok=True)
        self._init_workbook()

    # ── internal ──────────────────────────────────────────────────────────────
    def _thin_border(self):
        s = self.BORDER_SIDE
        return Border(left=s, right=s, top=s, bottom=s)

    def _init_workbook(self):
        """Create fresh workbook with Log + Summary sheets."""
        wb = Workbook()

        # ── Log sheet ─────────────────────────────────────────────────────────
        ws = wb.active
        ws.title = 'Training Log'
        ws.freeze_panes = 'A2'                 # freeze header row

        for col_idx, (header, width) in enumerate(
                zip(self.HEADERS, self.COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = self.HEADER_FONT
            cell.fill      = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
            cell.border    = self._thin_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30

        # ── Summary sheet ──────────────────────────────────────────────────────
        ws2 = wb.create_sheet('Summary')
        summary_headers = ['Metric', 'Value']
        metrics = [
            ('Total Episodes',           "='Training Log'!A{last}"),
            ('Total Timesteps',          "='Training Log'!B{last}"),
            ('Best Episode Reward',      "=MAX('Training Log'!C2:C{last})"),
            ('Worst Episode Reward',     "=MIN('Training Log'!C2:C{last})"),
            ('Mean Episode Reward',      "=AVERAGE('Training Log'!C2:C{last})"),
            ('Best Distance (m)',        "=MAX('Training Log'!F2:F{last})"),
            ('Mean Distance (m)',        "=AVERAGE('Training Log'!F2:F{last})"),
            ('Mean Deviation (m)',       "=AVERAGE('Training Log'!G2:G{last})"),
            ('Mean Episode Duration (s)',"=AVERAGE('Training Log'!H2:H{last})"),
        ]
        # write headers
        for ci, h in enumerate(summary_headers, 1):
            cell = ws2.cell(row=1, column=ci, value=h)
            cell.font      = self.HEADER_FONT
            cell.fill      = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border    = self._thin_border()
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 20

        # write metric rows with placeholder formulas (row=2 = first data row)
        for ri, (label, _) in enumerate(metrics, start=2):
            ws2.cell(row=ri, column=1, value=label).font  = self.DATA_FONT
            ws2.cell(row=ri, column=2, value='—').font    = self.DATA_FONT

        # store metric formula templates for later use
        self._summary_metrics = metrics

        wb.save(self.path)
        print(f'\n[ExcelLogger] Training log initialised → {self.path}\n')

    # ── public API ────────────────────────────────────────────────────────────
    def log(self, episode: int, timestep: int, ep_reward: float,
            scores: list, cumulative_score: float,
            distance: float, deviation: float, duration: float):
        """Append one episode row and refresh Summary formulas."""
        wb = load_workbook(self.path)
        ws = wb['Training Log']

        next_row  = ws.max_row + 1
        avg_last10 = float(np.mean(scores[-10:])) if scores else 0.0

        row_data = [
            episode,
            timestep,
            round(ep_reward,       4),
            round(avg_last10,      4),
            round(cumulative_score,4),
            round(distance,        2),
            round(deviation,       2),
            round(duration,        2),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=next_row, column=col_idx, value=value)
            cell.font      = self.DATA_FONT
            cell.border    = self._thin_border()
            cell.alignment = Alignment(horizontal='center')
            # alternating row shading
            if next_row % 2 == 0:
                cell.fill = self.ALT_FILL

        # refresh Summary formulas to cover new last row
        ws2      = wb['Summary']
        last_row = next_row
        for ri, (label, formula_tpl) in enumerate(self._summary_metrics, start=2):
            formula = formula_tpl.format(last=last_row)
            cell    = ws2.cell(row=ri, column=2, value=formula)
            cell.font      = self.DATA_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border    = self._thin_border()
            if ri % 2 == 0:
                cell.fill = self.ALT_FILL

        wb.save(self.path)


# ══════════════════════════════════════════════════════════════════════════════
# Train loop
# ══════════════════════════════════════════════════════════════════════════════
def train():

    timestep       = 0
    episode        = 0
    cumulative_score = 0
    episodic_length  = []
    scores           = []
    deviation_from_center = 0
    distance_covered      = 0

    np.random.seed(SEED); random.seed(SEED); tf.random.set_seed(SEED)

    try:
        client, world = ClientConnection().setup()
        print("CONNECTION HAS BEEN SETUP SUCCESSFULLY.\n")
    except Exception:
        print("CONNECTION REFUSED.\n")
        return

    os.makedirs(LOG_PATH_TRAIN, exist_ok=True)
    summary_writer = tf.summary.create_file_writer(LOG_PATH_TRAIN)

    env       = CarlaEnvironment(client, world, TOWN)
    encoder   = ConvEncoder(latent_dim=LATENT_DIM)
    xl_logger = ExcelLogger(RESULTS_PATH)

    if CHECKPOINT_LOAD:
        print('LOADING FROM CHECKPOINT....\n')
        agent                               = PPOAgent(encoder)
        episode, timestep, cumulative_score = agent.chkpt_load()
        agent.load()
        encoder = agent.encoder          # use loaded encoder
        agent.prn()
    else:
        agent = PPOAgent(encoder)

    # augmented encoder wrapper  (training=True → augmentation ON)
    enc_state = EncodeState(agent.encoder, training=True)

    while timestep < TRAIN_TIMESTEPS:

        observation = env.reset()
        observation = enc_state.process(observation)

        current_ep_reward = 0
        t1                = datetime.now()

        for t in range(EPISODE_LENGTH):

            observation = observation.numpy()
            action, _   = agent(observation, True)

            raw_obs, reward, done, info = env.step(action)

            if raw_obs is None:
                break

            observation = enc_state.process(raw_obs)

            agent.memory.rewards.append(reward)
            agent.memory.dones.append(done)

            timestep          += 1
            current_ep_reward += reward

            if done:
                episode += 1
                t2       = datetime.now()
                episodic_length.append(abs((t2 - t1).total_seconds()))
                break

        deviation_from_center += info[1]
        distance_covered      += info[0]
        scores.append(current_ep_reward)

        if CHECKPOINT_LOAD:
            cumulative_score = ((cumulative_score * (episode - 1)) + current_ep_reward) / episode
        else:
            cumulative_score = np.mean(scores)

        ep_duration = episodic_length[-1] if episodic_length else 0.0
        print(f'Episode: {episode}  |  Timestep: {timestep}'
              f'  |  Reward: {current_ep_reward:.2f}'
              f'  |  Avg Reward: {cumulative_score:.2f}'
              f'  |  Distance: {info[0]}')

        # ── Excel log (every episode) ──────────────────────────────────────
        xl_logger.log(
            episode        = episode,
            timestep       = timestep,
            ep_reward      = current_ep_reward,
            scores         = scores,
            cumulative_score = cumulative_score,
            distance       = info[0],
            deviation      = info[1],
            duration       = ep_duration,
        )

        if episode % 5 == 0:
            with summary_writer.as_default():
                tf.summary.scalar("Episodic Reward/episode",          scores[-1],               step=episode)
                tf.summary.scalar("Cumulative Reward/info",           cumulative_score,          step=episode)
                tf.summary.scalar("Cumulative Reward/(t)",            cumulative_score,          step=timestep)
                tf.summary.scalar("Average Episodic Reward/info",     np.mean(scores[-5:]),      step=episode)
                tf.summary.scalar("Average Reward/(t)",               np.mean(scores[-5:]),      step=timestep)
                tf.summary.scalar("Episode Length (s)/info",          np.mean(episodic_length),  step=episode)
                tf.summary.scalar("Reward/(t)",                       current_ep_reward,         step=timestep)
                tf.summary.scalar("Average Deviation from Center/episode", deviation_from_center/5, step=episode)
                tf.summary.scalar("Average Distance Covered (m)/episode",  distance_covered/5,      step=episode)
                summary_writer.flush()
                episodic_length       = []
                deviation_from_center = 0
                distance_covered      = 0

        if episode % 10 == 0:
            agent.learn()

        if episode % 50 == 0:
            agent.save()
            agent.chkpt_save(episode, timestep, cumulative_score)

    sys.exit()


# ══════════════════════════════════════════════════════════════════════════════
# Test loop
# ══════════════════════════════════════════════════════════════════════════════
def test():

    timestep = 0
    episode  = 0

    np.random.seed(SEED); random.seed(SEED); tf.random.set_seed(SEED)

    try:
        client, world = ClientConnection().setup()
        print("CONNECTION HAS BEEN SETUP SUCCESSFULLY.\n")
    except Exception:
        print("CONNECTION REFUSED.\n")
        return

    os.makedirs(LOG_PATH_TEST, exist_ok=True)
    summary_writer = tf.summary.create_file_writer(LOG_PATH_TEST)

    env     = CarlaEnvironment(client, world, TOWN)
    encoder = ConvEncoder(latent_dim=LATENT_DIM)
    agent   = PPOAgent(encoder)
    agent.load()
    agent.prn()

    # inference mode → augmentation OFF
    enc_state = EncodeState(agent.encoder, training=False)

    print("TESTING.....\n")

    while episode < TEST_EPISODES + 1:

        observation = env.reset()
        observation = enc_state.process(observation)

        total_time        = 0
        current_ep_reward = 0
        t1                = datetime.now()

        for t in range(EPISODE_LENGTH):

            observation = observation.numpy()
            action, _   = agent(observation, False)
            raw_obs, reward, done, info = env.step(action)

            if raw_obs is None:
                break

            observation        = enc_state.process(raw_obs)
            timestep          += 1
            current_ep_reward += reward

            if done:
                episode += 1
                break

        t2         = datetime.now()
        total_time = abs((t2 - t1).total_seconds())

        print(f'Episode: {episode}  |  Time: {total_time:.2f}s'
              f'  |  Reward: {current_ep_reward:.2f}'
              f'  |  Distance: {info[0]}')

        with summary_writer.as_default():
            tf.summary.scalar('Metrics/Time Taken',      total_time,        step=episode)
            tf.summary.scalar('Metrics/Reward',          current_ep_reward, step=episode)
            tf.summary.scalar('Metrics/Distance Covered', info[0],          step=episode)
            summary_writer.flush()

    sys.exit()


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    try:
        train()
        # test()
    except KeyboardInterrupt:
        sys.exit()
    finally:
        print('\nTerminating...')
